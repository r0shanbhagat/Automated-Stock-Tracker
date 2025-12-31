import yfinance as yf
import pandas as pd
import time
import json
from datetime import datetime
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import PatternFill
from datetime import timedelta
import os

# Ensure output folders exist
os.makedirs("previousdata", exist_ok=True)
os.makedirs("result", exist_ok=True)


# Normalize symbol by removing '.NS' suffix
def normalize_symbol(sym):
    return sym.replace(".NS", "") if sym else sym


# Load old trends for comparison (normalized symbol keys)
try:
    yesterday_str = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
    yesterday_filename = f'previousdata/previous_trends_{yesterday_str}.json'
    with open(yesterday_filename, 'r') as f:
        old_trends_raw = json.load(f)
        old_trends = {normalize_symbol(k): v for k, v in old_trends_raw.items()}
except FileNotFoundError:
    old_trends = {}

# Dictionary to store trends of current run (normalized symbol keys)
previous_trends = {}

# Create filename with date: previous_trends_YYYY-MM-DD.json inside "previousdata"
date_str = datetime.now().strftime('%Y-%m-%d')
filename = f'previousdata/previous_trends_{date_str}.json'


def save_previous_trends():
    normalized_save = {normalize_symbol(k): v for k, v in previous_trends.items()}
    with open(filename, 'w') as f:
        json.dump(normalized_save, f)


def compute_trend(ytd_price_perct):
    if ytd_price_perct is None:
        return "Unknown"
    if ytd_price_perct > 10:
        return "Bullish"
    elif ytd_price_perct < -10:
        return "Bearish"
    else:
        return "Neutral"


def get_returns_yahoo(symbol):
    try:
        ticker = yf.Ticker(symbol)
        hist = ticker.history(period="1y", interval="1d")
        if hist.empty:
            print(f"Yahoo returned empty for {symbol}")
            return None

        hist.index = hist.index.tz_localize(None)
        current = hist['Close'].iloc[-1]

        def pct(days):
            if len(hist) > days:
                return (hist['Close'].iloc[-1] / hist['Close'].iloc[-2] - 1) * 100 if days == 1 else \
                    (current / hist['Close'].iloc[-days] - 1) * 100
            return None

        start_year = datetime(datetime.today().year, 1, 1)
        ytd_data = hist[hist.index >= start_year]
        ytd_price = ytd_data['Close'].iloc[0] if not ytd_data.empty else None
        ytd_price_perct = round((current / ytd_price - 1) * 100, 2) if ytd_price else None

        qrt_price_perct = round(pct(63), 2) if pct(63) is not None else None

        return {
            "Current Price (₹)": round(float(current), 2),
            "1D %": round(pct(1), 2) if pct(1) else None,
            "1W %": round(pct(5), 2) if pct(5) else None,
            "2W %": round(pct(10), 2) if pct(10) else None,
            "1M %": round(pct(21), 2) if pct(21) else None,
            "3M %": qrt_price_perct,
            "6M %": round(pct(126), 2) if pct(126) else None,
            "YTD %": ytd_price_perct,
            "Trend": compute_trend(qrt_price_perct) if qrt_price_perct is not None else None,
            "Last Updated": datetime.now().strftime("%d-%m-%Y %H:%M:%S")
        }
    except Exception as e:
        print(f"Error fetching from Yahoo for {symbol} → {e}")
        return None


def detect_trend_change(symbol, current_trend):
    norm_symbol = normalize_symbol(symbol)
    previous_trend = previous_trends.get(norm_symbol)
    if previous_trend is None:
        previous_trends[norm_symbol] = current_trend
        return
    if previous_trend and current_trend and previous_trend != current_trend:
        print(f"Trend changed for {norm_symbol}: {previous_trend} → {current_trend}")
    previous_trends[norm_symbol] = current_trend


# Example Stock Mapping (complete with your 87 stocks)
stocks = {
    "Marksans Pharma Ltd": "MARKSANS.NS",
    "Jindal Poly Films Ltd": "JINDALPOLY.NS"
}

# "Oriental Rail Infrastructure Ltd: 531859
#"Patels Airtemp (India) Ltd": "517417",

# Master Loop
results = {}
for name, symbol in stocks.items():
    print(f"\nFetching {name} ({symbol})...")
    res = get_returns_yahoo(symbol)
    if res:
        detect_trend_change(symbol, res.get("Trend"))
    else:
        res = {"Error": "Data not found", "Trend": None}
    results[name] = {"Symbol": normalize_symbol(symbol), **res}
    time.sleep(0.5)

# Export results to Excel inside "result" folder
excelName = f"result/Stock-List_{date_str}.xlsx"
df = pd.DataFrame(results).T
df.to_excel(excelName)

wb = load_workbook(excelName)
ws = wb.active

# Find 'Trend' column index and rename header
trend_col_idx = None
for idx, cell in enumerate(ws[1], start=1):
    if cell.value == "Trend":
        trend_col_idx = idx
        cell.value = "Current Trend"
        break

# Insert "Trend Change" column
last_col = ws.max_column
trend_change_col_idx = last_col
ws.insert_cols(trend_change_col_idx)
ws.cell(row=1, column=trend_change_col_idx, value="Trend Change")

# Define fill colors
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
trend_fill_green = green_fill
trend_fill_red = red_fill

# Apply formatting and messages
for row in range(2, ws.max_row + 1):
    symbol = ws.cell(row=row, column=2).value  # Symbol assumed second column
    norm_symbol = normalize_symbol(symbol)
    trend_cell = ws.cell(row=row, column=trend_col_idx)
    trend_value = trend_cell.value

    if trend_value == "Bullish":
        trend_cell.fill = trend_fill_green
    elif trend_value == "Bearish":
        trend_cell.fill = trend_fill_red

    trend_change_cell = ws.cell(row=row, column=trend_change_col_idx)
    previous_trend = old_trends.get(norm_symbol)
    update_msg = ""
    fill_to_apply = None

    if previous_trend and trend_value and previous_trend != trend_value:
        update_msg = f"Trend changed from {previous_trend} to {trend_value}"
        if previous_trend == "Bearish" and trend_value == "Bullish":
            fill_to_apply = green_fill
        elif previous_trend == "Bullish" and trend_value == "Bearish":
            fill_to_apply = red_fill

    trend_change_cell.value = update_msg
    if fill_to_apply:
        trend_change_cell.fill = fill_to_apply

wb.save(excelName)

# Save trends for next run
save_previous_trends()

print(f"\n✅ {excelName} created with 'Current Trend' and 'Trend Change' colored columns!")
